from datetime import date

from openpyxl import load_workbook

from jobmatch_worker.exports.excel import build_excel_bytes, build_excel_workbook
from jobmatch_worker.exports.models import CandidateSummary, ExportRow, SearchCriteria


def make_row(**overrides: object) -> ExportRow:
    base = {
        "job_title": "Senior Backend Engineer",
        "company": "PT Contoh Nusantara",
        "location": "Jakarta",
        "region": "indonesia",
        "work_mode": "hybrid",
        "employment_type": "full-time",
        "salary_min": 15000000,
        "salary_max": 25000000,
        "currency": "IDR",
        "published_at": date(2026, 8, 10),
        "overall_score": 92.0,
        "skills_score": 95.0,
        "experience_score": 88.0,
        "education_score": 90.0,
        "location_score": 85.0,
        "seniority_score": 91.0,
        "language_score": 80.0,
        "verdict": "strong_match",
        "strengths": ["Python", "PostgreSQL"],
        "gaps": ["Kubernetes"],
        "critical_gaps": [],
        "recommendations": ["Add a project using Kubernetes."],
        "source_name": "jobstreet",
        "original_url": "https://example.com/jobs/42",
    }
    base.update(overrides)
    return ExportRow(**base)


def make_candidate() -> CandidateSummary:
    return CandidateSummary(
        name="Aulia Rahman",
        headline="Backend Engineer, 5 years",
        skills=["Python", "FastAPI", "PostgreSQL"],
        years_experience=5.0,
        location="Jakarta, Indonesia",
        languages=["Indonesian", "English"],
        education=["S1 Computer Science"],
    )


def make_criteria() -> SearchCriteria:
    return SearchCriteria(
        scope="best_and_strong",
        region=["indonesia"],
        work_mode=["hybrid"],
        min_score=80,
        status=["new", "saved"],
        date_from=date(2026, 8, 1),
        date_to=date(2026, 8, 31),
    )


def test_workbook_has_exact_sheets() -> None:
    wb = build_excel_workbook([make_row()], make_candidate(), make_criteria())
    assert wb.sheetnames == ["Job Matches", "Candidate Profile", "Search Criteria"]


def test_workbook_contains_only_filtered_rows() -> None:
    rows = [make_row(job_title="A"), make_row(job_title="B")]
    wb = build_excel_workbook(rows, make_candidate(), make_criteria())
    ws = wb["Job Matches"]
    assert ws.max_row == 3
    assert ws.cell(row=2, column=1).value == "A"
    assert ws.cell(row=3, column=1).value == "B"


def test_job_url_cells_are_hyperlinks() -> None:
    rows = [make_row(original_url="https://example.com/jobs/42")]
    wb = build_excel_workbook(rows, make_candidate(), make_criteria())
    ws = wb["Job Matches"]
    url_cell = ws.cell(row=2, column=24)
    assert url_cell.value == "https://example.com/jobs/42"
    assert url_cell.hyperlink is not None
    assert url_cell.hyperlink.target == "https://example.com/jobs/42"


def test_score_columns_are_numeric() -> None:
    wb = build_excel_workbook([make_row()], make_candidate(), make_criteria())
    ws = wb["Job Matches"]
    overall = ws.cell(row=2, column=11).value
    skills = ws.cell(row=2, column=12).value
    assert isinstance(overall, (int, float))
    assert isinstance(skills, (int, float))
    assert overall == 92.0
    assert skills == 95.0


def test_array_fields_serialized_with_newlines() -> None:
    wb = build_excel_workbook([make_row()], make_candidate(), make_criteria())
    ws = wb["Job Matches"]
    strengths = ws.cell(row=2, column=19).value
    assert strengths == "Python\nPostgreSQL"


def test_frozen_header_and_autofilter_present() -> None:
    wb = build_excel_workbook([make_row()], make_candidate(), make_criteria())
    ws = wb["Job Matches"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:X2"


def test_bytes_output_round_trips() -> None:
    blob = build_excel_bytes([make_row()], make_candidate(), make_criteria())
    assert blob[:2] == b"PK"
    wb = load_workbook(__import__("io").BytesIO(blob))
    assert wb.sheetnames == ["Job Matches", "Candidate Profile", "Search Criteria"]


def test_candidate_and_criteria_sheets_contain_summary() -> None:
    wb = build_excel_workbook([make_row()], make_candidate(), make_criteria())
    cp = wb["Candidate Profile"]
    values = {cp.cell(row=r, column=1).value: cp.cell(row=r, column=2).value for r in range(1, 8)}
    assert values["Name"] == "Aulia Rahman"
    assert values["Years of Experience"] == "5.0"
    assert "Python\nFastAPI\nPostgreSQL" in values["Skills"]
    sc = wb["Search Criteria"]
    criteria = {sc.cell(row=r, column=1).value: sc.cell(row=r, column=2).value for r in range(1, 8)}
    assert criteria["Scope"] == "best_and_strong"
    assert criteria["Minimum Score"] == "80"
    assert criteria["Region"] == "indonesia"