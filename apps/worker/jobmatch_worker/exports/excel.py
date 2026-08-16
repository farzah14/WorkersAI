from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import CandidateSummary, ExportRow, SearchCriteria

COLUMNS = [
    "Job Title",
    "Company",
    "Location",
    "Region",
    "Work Mode",
    "Employment Type",
    "Salary Min",
    "Salary Max",
    "Currency",
    "Published Date",
    "Overall Match",
    "Skills",
    "Experience",
    "Education",
    "Location Score",
    "Seniority",
    "Language",
    "Verdict",
    "Strengths",
    "Gaps",
    "Critical Gaps",
    "AI Recommendation",
    "Source",
    "Job URL",
]

_URL_COLUMN = len(COLUMNS)


def _serialize(values: list[str]) -> str:
    return "\n".join(values)


def _fmt_date(value) -> str:
    return value.isoformat() if value is not None else ""


def _row_values(row: ExportRow) -> list[str | int | float | None]:
    return [
        row.job_title,
        row.company,
        row.location or "",
        row.region or "",
        row.work_mode or "",
        row.employment_type or "",
        row.salary_min,
        row.salary_max,
        row.currency or "",
        _fmt_date(row.published_at),
        row.overall_score,
        row.skills_score,
        row.experience_score,
        row.education_score,
        row.location_score,
        row.seniority_score,
        row.language_score,
        row.verdict,
        _serialize(row.strengths),
        _serialize(row.gaps),
        _serialize(row.critical_gaps),
        _serialize(row.recommendations),
        row.source_name,
        row.original_url,
    ]


def _write_label_value(ws, rows: list[tuple[str, str]]) -> None:
    for index, (label, value) in enumerate(rows, start=1):
        ws.cell(row=index, column=1, value=label).font = Font(bold=True)
        ws.cell(row=index, column=2, value=value)


def build_excel_workbook(
    rows: list[ExportRow],
    candidate: CandidateSummary,
    criteria: SearchCriteria,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Job Matches"

    for col, title in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col, value=title).font = Font(bold=True)

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(_row_values(row), start=1):
            ws.cell(row=r, column=c, value=value)
        url_cell = ws.cell(row=r, column=_URL_COLUMN, value=row.original_url)
        url_cell.hyperlink = row.original_url
        url_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(_URL_COLUMN)}{len(rows) + 1}"
    ws.column_dimensions[get_column_letter(_URL_COLUMN)].width = 48
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18

    cp = wb.create_sheet("Candidate Profile")
    _write_label_value(
        cp,
        [
            ("Name", candidate.name),
            ("Headline", candidate.headline),
            (
                "Years of Experience",
                str(candidate.years_experience) if candidate.years_experience is not None else "",
            ),
            ("Location", candidate.location),
            ("Skills", _serialize(candidate.skills)),
            ("Languages", _serialize(candidate.languages)),
            ("Education", _serialize(candidate.education)),
        ],
    )

    sc = wb.create_sheet("Search Criteria")
    _write_label_value(
        sc,
        [
            ("Scope", criteria.scope),
            ("Region", _serialize(criteria.region or [])),
            ("Work Mode", _serialize(criteria.work_mode or [])),
            ("Minimum Score", str(criteria.min_score) if criteria.min_score is not None else ""),
            ("Status", _serialize(criteria.status or [])),
            ("Date From", _fmt_date(criteria.date_from)),
            ("Date To", _fmt_date(criteria.date_to)),
        ],
    )

    return wb


def build_excel_bytes(
    rows: list[ExportRow],
    candidate: CandidateSummary,
    criteria: SearchCriteria,
) -> bytes:
    buffer = BytesIO()
    build_excel_workbook(rows, candidate, criteria).save(buffer)
    return buffer.getvalue()